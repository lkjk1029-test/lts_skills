#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
엑셀 보고서 생성 스크립트
웹 보안 분석 결과를 엑셀 파일로 생성
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference
import json
from datetime import datetime
import os
from typing import Dict, List, Any

class ExcelReportGenerator:
    """웹 보안 분석 결과 엑셀 보고서 생성기"""

    def __init__(self, analysis_results: Dict[str, Any]):
        self.analysis_results = analysis_results
        self.workbook = None
        self.current_row = 1

    def create_report(self, output_filename: str = None) -> str:
        """보고서 생성 메인 함수"""

        if output_filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"web_security_analysis_{timestamp}.xlsx"

        # 워크북 생성
        self.workbook = openpyxl.Workbook()

        # 기본 시트 삭제
        self.workbook.remove(self.workbook.active)

        # 각 시트 생성
        self._create_summary_sheet()
        self._create_basic_info_sheet()
        self._create_security_sheet()
        self._create_forms_sheet()
        self._create_navigation_sheet()
        self._create_storage_sheet()
        self._create_network_sheet()
        self._create_vulnerabilities_sheet()
        self._create_recommendations_sheet()

        # 파일 저장
        self.workbook.save(output_filename)
        print(f"Excel report created: {output_filename}")

        return output_filename

    def _create_summary_sheet(self):
        """요약 정보 시트 생성"""
        ws = self.workbook.create_sheet("요약 정보")

        # 제목
        self._add_title(ws, "웹 보안 분석 보고서 요약")

        # 기본 정보
        basic_info = self.analysis_results.get('basic_info', {})
        security_info = self.analysis_results.get('security', {})

        summary_data = [
            ["분석 대상 URL", basic_info.get('url', 'N/A')],
            ["사이트 제목", basic_info.get('title', 'N/A')],
            ["분석 시간", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["프로토콜", basic_info.get('protocol', 'N/A')],
            ["HTTPS 사용", "예" if security_info.get('isHTTPS') else "아니오"],
            ["Mixed Content", f"{security_info.get('totalMixedContent', 0)}개"],
            ["총 폼 수", len(self.analysis_results.get('forms', []))],
            ["내부 링크 수", self.analysis_results.get('navigation', {}).get('internalLinks', 0)],
            ["외부 링크 수", self.analysis_results.get('navigation', {}).get('externalLinks', 0)],
            ["콘솔 오류 수", self.analysis_results.get('console', {}).get('errorCount', 0)],
            ["콘솔 경고 수", self.analysis_results.get('console', {}).get('warningCount', 0)],
        ]

        self._add_table(ws, summary_data, start_col=1, start_row=self.current_row)

        # 취약점 요약
        self.current_row += len(summary_data) + 3
        self._add_subtitle(ws, "취약점 요약")

        vulnerabilities = self.analysis_results.get('vulnerabilities', {})

        vuln_data = [
            ["위험도", "개수", "비율"],
            ["높음 (High)", len(vulnerabilities.get('high', [])), ""],
            ["중간 (Medium)", len(vulnerabilities.get('medium', [])), ""],
            ["낮음 (Low)", len(vulnerabilities.get('low', [])), ""],
            ["총계", self._count_total_vulnerabilities(), "100%"]
        ]

        # 비율 계산
        total_vulns = self._count_total_vulnerabilities()
        if total_vulns > 0:
            vuln_data[1][2] = f"{len(vulnerabilities.get('high', []))/total_vulns*100:.1f}%"
            vuln_data[2][2] = f"{len(vulnerabilities.get('medium', []))/total_vulns*100:.1f}%"
            vuln_data[3][2] = f"{len(vulnerabilities.get('low', []))/total_vulns*100:.1f}%"

        self._add_table(ws, vuln_data, start_col=1, start_row=self.current_row)

        # 스타일 적용
        self._format_summary_sheet(ws)

    def _create_basic_info_sheet(self):
        """기본 정보 시트 생성"""
        ws = self.workbook.create_sheet("기본 정보")

        self._add_title(ws, "웹사이트 기본 정보")

        basic_info = self.analysis_results.get('basic_info', {})

        info_data = [
            ["항목", "값"],
            ["URL", basic_info.get('url', 'N/A')],
            ["도메인", basic_info.get('domain', 'N/A')],
            ["프로토콜", basic_info.get('protocol', 'N/A')],
            ["포트", basic_info.get('port', 'N/A')],
            ["경로", basic_info.get('path', 'N/A')],
            ["제목", basic_info.get('title', 'N/A')],
            ["언어", basic_info.get('language', 'N/A')],
            ["플랫폼", basic_info.get('platform', 'N/A')],
            ["User Agent", basic_info.get('userAgent', 'N/A')],
            ["문자셋", basic_info.get('characterSet', 'N/A')],
            ["레퍼러", basic_info.get('referrer', 'N/A')],
            ["쿠키 사용 가능", "예" if basic_info.get('cookieEnabled') else "아니오"],
            ["온라인 상태", "예" if basic_info.get('onLine') else "아니오"],
        ]

        self._add_table(ws, info_data)
        self._format_data_sheet(ws)

    def _create_security_sheet(self):
        """보안 분석 시트 생성"""
        ws = self.workbook.create_sheet("보안 분석")

        self._add_title(ws, "HTTPS 및 보안 설정 분석")

        security = self.analysis_results.get('security', {})

        security_data = [
            ["항목", "상태", "설명"],
            ["HTTPS 사용", "양호" if security.get('isHTTPS') else "취약",
             "HTTPS를 사용하면 데이터 전송이 암호화됩니다" if security.get('isHTTPS') else
             "HTTP를 사용하면 데이터가 평문으로 전송됩니다"],
            ["Mixed Content", "양호" if security.get('totalMixedContent', 0) == 0 else "취약",
             f"Mixed Content가 없습니다" if security.get('totalMixedContent', 0) == 0 else
             f"{security.get('totalMixedContent', 0)}개의 HTTP 리소스가 있습니다"],
        ]

        # Mixed Content 상세
        mixed_content = security.get('mixedContent', {})
        if mixed_content:
            security_data.extend([
                ["  - HTTP 이미지", mixed_content.get('httpImages', 0), "개"],
                ["  - HTTP 스크립트", mixed_content.get('httpScripts', 0), "개"],
                ["  - HTTP 스타일시트", mixed_content.get('httpStyles', 0), "개"],
                ["  - HTTP 아이프레임", mixed_content.get('httpIframes', 0), "개"],
            ])

        # CSP 확인
        self.current_row += len(security_data) + 2
        self._add_subtitle(ws, "Content Security Policy (CSP)")

        csp_info = security.get('cspMeta', "설정되지 않음")
        csp_data = [
            ["CSP 설정", csp_info],
            ["보안 등급", "양호" if csp_info and csp_info != "설정되지 않음" else "주의 필요"]
        ]

        self._add_table(ws, security_data)
        self._add_table(ws, csp_data, start_col=1, start_row=self.current_row)

        # 외부 리소스 분석
        if security.get('externalResources'):
            self.current_row += len(csp_data) + 2
            self._add_subtitle(ws, "외부 리소스 분석")

            external_data = [["태그", "URL", "외부 도메인", "HTTP 여부", "Integrity"]]
            for resource in security['externalResources'][:20]:  # 최대 20개만 표시
                external_data.append([
                    resource.get('tagName', ''),
                    resource.get('url', ''),
                    resource.get('isExternal', ''),
                    "예" if resource.get('isHTTP') else "아니오",
                    resource.get('integrity', '')
                ])

            self._add_table(ws, external_data)

        self._format_security_sheet(ws)

    def _create_forms_sheet(self):
        """폼 분석 시트 생성"""
        ws = self.workbook.create_sheet("폼 분석")

        self._add_title(ws, "웹 폼 및 입력 필드 분석")

        forms = self.analysis_results.get('forms', [])

        if not forms:
            ws.cell(row=self.current_row, column=1, value="폼이 발견되지 않았습니다.")
            return

        for i, form in enumerate(forms):
            if i > 0:  # 첫 폼이 아니라면 간격 추가
                self.current_row += 2

            # 폼 기본 정보
            self._add_subtitle(ws, f"폼 #{i+1}: {form.get('action', 'N/A')}")

            form_data = [
                ["속성", "값"],
                ["Action", form.get('action', 'N/A')],
                ["Method", form.get('method', 'N/A')],
                ["ID", form.get('id', 'N/A')],
                ["Class", form.get('className', 'N/A')],
                ["Encoding", form.get('enctype', 'N/A')],
                ["CSRF 토큰", form.get('csrfToken', '없음')],
            ]

            self._add_table(ws, form_data)
            self.current_row += len(form_data) + 1

            # 입력 필드 정보
            fields = form.get('fields', [])
            if fields:
                self._add_subtitle(ws, f"입력 필드 ({len(fields)}개)")

                field_data = [["타입", "이름", "ID", "필수", "자동완성", "최대길이", "보안 관련"]]
                for field in fields:
                    security_notes = []
                    if field.get('isPassword'):
                        security_notes.append("비밀번호")
                    if field.get('isEmail'):
                        security_notes.append("이메일")
                    if field.get('isFile'):
                        security_notes.append("파일업로드")
                    if field.get('isHidden'):
                        security_notes.append("숨김필드")

                    field_data.append([
                        field.get('type', ''),
                        field.get('name', ''),
                        field.get('id', ''),
                        "예" if field.get('required') else "아니오",
                        field.get('autocomplete', ''),
                        field.get('maxlength', '') if field.get('maxlength') != -1 else '제한없음',
                        ", ".join(security_notes) if security_notes else "-"
                    ])

                self._add_table(ws, field_data)

            # 잠재적 취약점
            vulnerabilities = form.get('potentialVulnerabilities', [])
            if vulnerabilities:
                self.current_row += len(field_data) + 1
                self._add_subtitle(ws, "잠재적 취약점")

                vuln_data = [["유형", "필드", "위험도"]]
                for vuln in vulnerabilities:
                    vuln_data.append([
                        vuln.get('type', ''),
                        vuln.get('field', ''),
                        vuln.get('severity', '')
                    ])

                self._add_table(ws, vuln_data)

        self._format_forms_sheet(ws)

    def _create_navigation_sheet(self):
        """내비게이션 분석 시트 생성"""
        ws = self.workbook.create_sheet("내비게이션 분석")

        self._add_title(ws, "링크 및 내비게이션 구조 분석")

        nav = self.analysis_results.get('navigation', {})

        # 전체 링크 통계
        stats_data = [
            ["항목", "개수"],
            ["전체 링크", nav.get('totalLinks', 0)],
            ["내부 링크", nav.get('internalLinks', 0)],
            ["외부 링크", nav.get('externalLinks', 0)],
            ["앵커 링크", nav.get('anchorLinks', 0)],
            ["JavaScript 링크", nav.get('javascriptLinks', 0)],
            ["메일to 링크", nav.get('mailtoLinks', 0)],
            ["전화 링크", nav.get('telLinks', 0)],
        ]

        self._add_table(ws, stats_data)

        # 내비게이션 메뉴 구조
        nav_menus = nav.get('navMenus', [])
        if nav_menus:
            self.current_row += len(stats_data) + 2
            self._add_subtitle(ws, "내비게이션 메뉴 구조")

            for i, menu in enumerate(nav_menus[:5]):  # 최대 5개 메뉴만 표시
                if i > 0:
                    self.current_row += 1

                self._add_subtitle(ws, f"메뉴 #{i+1}: {menu.get('className', 'N/A')}")

                menu_data = [["텍스트", "URL", "내부 링크 여부"]]
                for link in menu.get('links', [])[:10]:  # 메뉴당 최대 10개 링크
                    menu_data.append([
                        link.get('text', ''),
                        link.get('href', ''),
                        "예" if link.get('isInternal') else "아니오"
                    ])

                self._add_table(ws, menu_data)

        # 외부 링크 상세
        all_links = nav.get('allLinks', [])
        external_links = [link for link in all_links if link.get('isExternal')]

        if external_links:
            self.current_row += 2
            self._add_subtitle(ws, "외부 링크 목록 (최대 20개)")

            ext_data = [["텍스트", "URL", "Target", "Rel"]]
            for link in external_links[:20]:
                ext_data.append([
                    link.get('text', ''),
                    link.get('href', ''),
                    link.get('target', ''),
                    link.get('rel', '')
                ])

            self._add_table(ws, ext_data)

        self._format_navigation_sheet(ws)

    def _create_storage_sheet(self):
        """스토리지 분석 시트 생성"""
        ws = self.workbook.create_sheet("스토리지 분석")

        self._add_title(ws, "쿠키 및 웹 스토리지 분석")

        storage = self.analysis_results.get('storage', {})

        # 쿠키 정보
        cookies = storage.get('cookies', {})
        self._add_subtitle(ws, f"쿠키 ({cookies.get('count', 0)}개)")

        cookie_data = [["이름", "크기", "도메인"]]
        for cookie in cookies.get('items', [])[:20]:  # 최대 20개 쿠키
            cookie_data.append([
                cookie.get('name', ''),
                cookie.get('size', 0),
                cookie.get('domain', '')
            ])

        if cookie_data:
            self._add_table(ws, cookie_data)

        # localStorage
        self.current_row += len(cookie_data) + 2
        local_storage = storage.get('localStorage', {})
        self._add_subtitle(ws, f"Local Storage ({local_storage.get('count', 0)}개, {local_storage.get('totalSize', 0)} bytes)")

        local_data = [["키", "크기", "값 (일부)"]]
        items_list = list(local_storage.get('items', {}).items())[:15]  # 최대 15개 항목
        for key, info in items_list:
            value_preview = str(info.get('value', ''))[:50] + "..." if len(str(info.get('value', ''))) > 50 else str(info.get('value', ''))
            local_data.append([key, info.get('size', 0), value_preview])

        if local_data:
            self._add_table(ws, local_data)

        # sessionStorage
        self.current_row += len(local_data) + 2
        session_storage = storage.get('sessionStorage', {})
        self._add_subtitle(ws, f"Session Storage ({session_storage.get('count', 0)}개, {session_storage.get('totalSize', 0)} bytes)")

        session_data = [["키", "크기", "값 (일부)"]]
        items_list = list(session_storage.get('items', {}).items())[:15]  # 최대 15개 항목
        for key, info in items_list:
            value_preview = str(info.get('value', ''))[:50] + "..." if len(str(info.get('value', ''))) > 50 else str(info.get('value', ''))
            session_data.append([key, info.get('size', 0), value_preview])

        if session_data:
            self._add_table(ws, session_data)

        # 민감정보 저장 확인
        sensitive_data = storage.get('sensitiveData', [])
        if sensitive_data:
            self.current_row += len(session_data) + 2
            self._add_subtitle(ws, "⚠️ 민감정보 저장 현황")

            sensitive_table = [["저장소", "키", "패턴", "크기"]]
            for item in sensitive_data:
                sensitive_table.append([
                    item.get('container', ''),
                    item.get('key', ''),
                    item.get('pattern', ''),
                    item.get('size', 0)
                ])

            self._add_table(ws, sensitive_table)

        self._format_storage_sheet(ws)

    def _create_network_sheet(self):
        """네트워크 분석 시트 생성"""
        ws = self.workbook.create_sheet("네트워크 분석")

        self._add_title(ws, "네트워크 요청 분석")

        network = self.analysis_results.get('network', {})

        # 네트워크 통계
        stats_data = [
            ["항목", "개수"],
            ["총 요청", network.get('total', 0)],
            ["HTTPS 요청", network.get('httpsRequests', 0)],
            ["HTTP 요청", network.get('httpRequests', 0)],
            ["내부 요청", network.get('internalRequests', 0)],
            ["외부 요청", network.get('externalRequests', 0)],
            ["API 엔드포인트", len(network.get('apiEndpoints', []))],
        ]

        self._add_table(ws, stats_data)

        # 요청 타입별 분석
        by_type = network.get('byType', {})
        if by_type:
            self.current_row += len(stats_data) + 2
            self._add_subtitle(ws, "요청 타입별 분석")

            type_data = [["타입", "개수"]]
            for req_type, count in by_type.items():
                type_data.append([req_type, count])

            self._add_table(ws, type_data)

        # API 엔드포인트
        api_endpoints = network.get('apiEndpoints', [])
        if api_endpoints:
            self.current_row += len(type_data) + 2
            self._add_subtitle(ws, "API 엔드포인트")

            api_data = [["URL", "메소드", "타입"]]
            for endpoint in api_endpoints[:20]:  # 최대 20개
                api_data.append([
                    endpoint.get('url', ''),
                    endpoint.get('method', ''),
                    endpoint.get('type', '')
                ])

            self._add_table(ws, api_data)

        # 잠재적 취약점
        potential_vulns = network.get('potentialVulnerabilities', [])
        if potential_vulns:
            self.current_row += len(api_data) + 2
            self._add_subtitle(ws, "⚠️ 잠재적 네트워크 취약점")

            vuln_data = [["유형", "URL", "메소드"]]
            for vuln in potential_vulns:
                vuln_data.append([
                    vuln.get('type', ''),
                    vuln.get('url', ''),
                    vuln.get('method', '')
                ])

            self._add_table(ws, vuln_data)

        self._format_network_sheet(ws)

    def _create_vulnerabilities_sheet(self):
        """취약점 상세 시트 생성"""
        ws = self.workbook.create_sheet("취약점 상세")

        self._add_title(ws, "보안 취약점 상세 분석")

        vulnerabilities = self.analysis_results.get('vulnerabilities', {})

        # 높음 위험도
        high_vulns = vulnerabilities.get('high', [])
        if high_vulns:
            self._add_subtitle(ws, "🔴 높음 위험도 취약점")
            high_data = [["#", "취약점", "설명"]]
            for i, vuln in enumerate(high_vulns, 1):
                high_data.append([i, vuln, self._get_vulnerability_description(vuln, 'high')])
            self._add_table(ws, high_data)

        # 중간 위험도
        medium_vulns = vulnerabilities.get('medium', [])
        if medium_vulns:
            self.current_row += len(high_data) + 2 if high_vulns else 2
            self._add_subtitle(ws, "🟡 중간 위험도 취약점")
            medium_data = [["#", "취약점", "설명"]]
            for i, vuln in enumerate(medium_vulns, 1):
                medium_data.append([i, vuln, self._get_vulnerability_description(vuln, 'medium')])
            self._add_table(ws, medium_data)

        # 낮음 위험도
        low_vulns = vulnerabilities.get('low', [])
        if low_vulns:
            self.current_row += len(medium_data) + 2 if medium_vulns else 2
            self._add_subtitle(ws, "🟢 낮음 위험도 취약점")
            low_data = [["#", "취약점", "설명"]]
            for i, vuln in enumerate(low_vulns, 1):
                low_data.append([i, vuln, self._get_vulnerability_description(vuln, 'low')])
            self._add_table(ws, low_data)

        if not any([high_vulns, medium_vulns, low_vulns]):
            ws.cell(row=self.current_row, column=1, value="발견된 취약점이 없습니다.")

        self._format_vulnerabilities_sheet(ws)

    def _create_recommendations_sheet(self):
        """권장 사항 시트 생성"""
        ws = self.workbook.create_sheet("권장 사항")

        self._add_title(ws, "보안 강화 권장 사항")

        recommendations = self._generate_recommendations()

        if recommendations:
            rec_data = [["우선순위", "권장 사항", "적용 방법", "예상 효과"]]

            for rec in recommendations:
                rec_data.append([
                    rec.get('priority', ''),
                    rec.get('recommendation', ''),
                    rec.get('how_to', ''),
                    rec.get('benefit', '')
                ])

            self._add_table(ws, rec_data)
        else:
            ws.cell(row=self.current_row, column=1, value="현재 상태가 양호하여 추가 권장 사항이 없습니다.")

        self._format_recommendations_sheet(ws)

    # 도우미 메소드들
    def _add_title(self, ws, title):
        """제목 추가"""
        cell = ws.cell(row=self.current_row, column=1, value=title)
        cell.font = Font(size=16, bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(size=16, bold=True, color="FFFFFF")
        ws.merge_cells(start_row=self.current_row, start_column=1, end_row=self.current_row, end_column=6)
        self.current_row += 2

    def _add_subtitle(self, ws, subtitle):
        """부제목 추가"""
        cell = ws.cell(row=self.current_row, column=1, value=subtitle)
        cell.font = Font(size=12, bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        ws.merge_cells(start_row=self.current_row, start_column=1, end_row=self.current_row, end_column=6)
        self.current_row += 2

    def _add_table(self, ws, data, start_col=1, start_row=None):
        """테이블 추가"""
        if start_row is None:
            start_row = self.current_row

        for r_idx, row in enumerate(data, 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=start_row + r_idx - 1, column=start_col + c_idx - 1, value=value)

                # 헤더 행 스타일
                if r_idx == 1:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                       top=Side(style='thin'), bottom=Side(style='thin'))
                else:
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                       top=Side(style='thin'), bottom=Side(style='thin'))

        # 열 너비 자동 조절
        for c_idx in range(1, len(data[0]) + 1):
            max_length = 0
            for row in data:
                if len(row) >= c_idx:
                    cell_length = len(str(row[c_idx - 1]))
                    if cell_length > max_length:
                        max_length = cell_length
            ws.column_dimensions[openpyxl.utils.get_column_letter(start_col + c_idx - 1)].width = min(max_length + 2, 50)

        self.current_row = start_row + len(data) + 1

    def _count_total_vulnerabilities(self):
        """총 취약점 수 계산"""
        vulnerabilities = self.analysis_results.get('vulnerabilities', {})
        return (len(vulnerabilities.get('high', [])) +
                len(vulnerabilities.get('medium', [])) +
                len(vulnerabilities.get('low', [])))

    def _get_vulnerability_description(self, vulnerability, severity):
        """취약점 설명 생성"""
        descriptions = {
            'high': {
                'HTTP 프로토콜 사용': '데이터가 암호화되지 않고 전송되어 중간자 공격에 취약합니다.',
                '비밀번호 전송에 GET 방식 사용': '비밀번호가 URL에 노출되어 브라우저 기록이나 로그에 남을 수 있습니다.',
                'Mixed Content': 'HTTPS 페이지에서 HTTP 리소스를 로드하여 보안이 취약해집니다.',
                '민감정보 URL 노출': 'API 키나 비밀번호 같은 민감정보가 URL에 노출됩니다.'
            },
            'medium': {
                '콘솔 오류': '시스템 내부 정보가 노출될 수 있습니다.',
                '민감정보 localStorage 저장': '클라이언트 측에 민감정보가 저장되어 XSS 공격에 취약합니다.',
                '인라인 스크립트에 민감정보': '소스 코드에 민감정보가 노출됩니다.'
            },
            'low': {
                '비밀번호 autocomplete disabled': '사용자 경험을 저해하고 강제적인 비밀번호 관리를 유발할 수 있습니다.',
                '디버깅 정보 노출 가능성': '개발 관련 정보가 노출될 수 있습니다.'
            }
        }

        return descriptions.get(severity, {}).get(vulnerability, '상세 설명 준비 중...')

    def _generate_recommendations(self):
        """권장 사항 생성"""
        recommendations = []

        # HTTPS 권장
        if not self.analysis_results.get('security', {}).get('isHTTPS'):
            recommendations.append({
                'priority': '높음',
                'recommendation': 'HTTPS 적용',
                'how_to': 'SSL/TLS 인증서 설치 및 모든 HTTP 요청을 HTTPS로 리다이렉트',
                'benefit': '데이터 전송 암호화 및 사용자 신뢰도 향상'
            })

        # Mixed Content 해결
        if self.analysis_results.get('security', {}).get('totalMixedContent', 0) > 0:
            recommendations.append({
                'priority': '중간',
                'recommendation': 'Mixed Content 제거',
                'how_to': '모든 HTTP 리소스를 HTTPS로 변경',
                'benefit': '브라우저 보안 경고 제거 및 데이터 무결성 보장'
            })

        # CSP 설정
        if not self.analysis_results.get('security', {}).get('cspMeta'):
            recommendations.append({
                'priority': '중간',
                'recommendation': 'Content Security Policy 설정',
                'how_to': 'CSP 헤더 또는 meta 태그를 통해 허용된 리소스 소스 지정',
                'benefit': 'XSS 및 인젝션 공격 방지'
            })

        # 폼 보안 강화
        forms = self.analysis_results.get('forms', [])
        for form in forms:
            if form.get('method') == 'GET' and any(field.get('isPassword') for field in form.get('fields', [])):
                recommendations.append({
                    'priority': '높음',
                    'recommendation': '폼 전송 방식을 POST로 변경',
                    'how_to': 'form 태그의 method 속성을 POST로 변경',
                    'benefit': '민감정보가 URL이나 서버 로그에 노출되지 않음'
                })
                break

        return recommendations

    # 스타일 적용 메소드들
    def _format_summary_sheet(self, ws):
        """요약 시트 스타일 적용"""
        pass  # 기본 테이블 스타일로 충분

    def _format_data_sheet(self, ws):
        """데이터 시트 스타일 적용"""
        pass

    def _format_security_sheet(self, ws):
        """보안 시트 스타일 적용"""
        for row in ws.iter_rows():
            for cell in row:
                if "취약" in str(cell.value):
                    cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                elif "양호" in str(cell.value):
                    cell.fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")

    def _format_forms_sheet(self, ws):
        """폼 시트 스타일 적용"""
        pass

    def _format_navigation_sheet(self, ws):
        """내비게이션 시트 스타일 적용"""
        pass

    def _format_storage_sheet(self, ws):
        """스토리지 시트 스타일 적용"""
        pass

    def _format_network_sheet(self, ws):
        """네트워크 시트 스타일 적용"""
        pass

    def _format_vulnerabilities_sheet(self, ws):
        """취약점 시트 스타일 적용"""
        for row in ws.iter_rows():
            for cell in row:
                if "🔴" in str(cell.value):
                    cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                elif "🟡" in str(cell.value):
                    cell.fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
                elif "🟢" in str(cell.value):
                    cell.fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")

    def _format_recommendations_sheet(self, ws):
        """권장 사항 시트 스타일 적용"""
        pass

def main():
    """메인 함수 - 테스트용"""
    # 테스트 데이터
    test_data = {
        'basic_info': {
            'url': 'https://example.com',
            'title': 'Example Website',
            'domain': 'example.com',
            'protocol': 'https:',
            'language': 'ko'
        },
        'security': {
            'isHTTPS': True,
            'totalMixedContent': 2,
            'cspMeta': None
        },
        'forms': [
            {
                'action': '/login',
                'method': 'POST',
                'fields': [
                    {'name': 'username', 'type': 'text', 'required': True},
                    {'name': 'password', 'type': 'password', 'required': True}
                ],
                'potentialVulnerabilities': []
            }
        ],
        'navigation': {
            'totalLinks': 15,
            'internalLinks': 12,
            'externalLinks': 3
        },
        'storage': {
            'cookies': {'count': 2, 'items': [{'name': 'sessionid', 'size': 32}]},
            'localStorage': {'count': 1, 'totalSize': 100},
            'sessionStorage': {'count': 0, 'totalSize': 0}
        },
        'network': {
            'total': 25,
            'httpsRequests': 23,
            'httpRequests': 2
        },
        'vulnerabilities': {
            'high': ['Mixed Content 발견'],
            'medium': ['콘솔 오류 발견'],
            'low': ['디버깅 정보 노출 가능성']
        }
    }

    generator = ExcelReportGenerator(test_data)
    output_file = generator.create_report("test_security_report.xlsx")
    print(f"테스트 보고서 생성: {output_file}")

if __name__ == "__main__":
    main()